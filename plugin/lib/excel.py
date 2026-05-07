"""Single-sheet xlsx report writer.

One row per candidate query. Columns match the analyzer output schema,
with two display conveniences:
  - bytes columns rendered as GB (decimal, /1e9, 2 decimals)
  - elapsed-time columns rendered as seconds (/1000, 2 decimals)

Records passed in still use the raw bytes/ms keys (`bytes_scanned`,
`elapsed_ms`) — the writer is the single point that knows about display
units, so callers don't need to convert.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Conversion constants. Decimal GB matches what most tools (Snowflake's own
# UI included, in casual context) call "GB"; rounded to 2 decimals.
_BYTES_PER_GB = 10 ** 9
_MS_PER_SEC = 1000
_GB_DECIMALS = 2
_SEC_DECIMALS = 2


def _bytes_to_gb(b: float | int | None) -> float | None:
    if b is None:
        return None
    return round(b / _BYTES_PER_GB, _GB_DECIMALS)


def _ms_to_sec(ms: float | int | None) -> float | None:
    if ms is None:
        return None
    return round(ms / _MS_PER_SEC, _SEC_DECIMALS)


# Each entry: (display_column_name, source_record_key, column_width_chars,
# converter_or_None). source_key may equal display_name when no conversion.
COLUMNS: list[tuple[str, str, int, callable | None]] = [
    ("query_id",              "query_id",              22, None),
    ("sql",                   "sql",                   60, None),
    ("elapsed_sec",           "elapsed_ms",            16, _ms_to_sec),
    ("gb_scanned",            "bytes_scanned",         16, _bytes_to_gb),
    ("estimated_credits",     "estimated_credits",     18, None),
    ("warehouse_name",        "warehouse_name",        18, None),
    ("warehouse_size",        "warehouse_size",        14, None),
    ("recommendation_type",   "recommendation_type",   22, None),
    ("recommendation_target", "recommendation_target", 28, None),
    ("rationale",             "rationale",             60, None),
]

SQL_TRUNCATE_LEN = 8000  # Excel cell limit is 32767; keep readable


def _truncate(s: str | None, n: int = SQL_TRUNCATE_LEN) -> str:
    if s is None:
        return ""
    if len(s) <= n:
        return s
    return s[: n - 1] + "…"


def write_report(records: list[dict], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "candidates"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    for col_idx, (display, _src, width, _conv) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=display)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = "A2"

    wrap_cols = {"sql", "rationale"}
    for row_idx, rec in enumerate(records, start=2):
        for col_idx, (display, source_key, _w, conv) in enumerate(COLUMNS, start=1):
            val = rec.get(source_key)
            if display == "sql":
                val = _truncate(val)
            elif conv is not None:
                val = conv(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if display in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(out_path)
