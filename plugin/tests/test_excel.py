"""Excel writer tests. Covers the unit conversions (bytes → GB, ms → sec)
and the basic round-trip through openpyxl. The writer is small enough that
the structural tests below are sufficient — heavy schema validation lives
in the smoke test in TESTING.md §3."""
from __future__ import annotations

import os
import tempfile

import pytest
from openpyxl import load_workbook

from lib.excel import COLUMNS, _bytes_to_gb, _ms_to_sec, write_report


# ---- conversion units ----

@pytest.mark.parametrize(
    "bytes_in,gb_out",
    [
        (None, None),
        (0, 0.0),
        (1_000_000_000, 1.0),
        (1_500_000_000, 1.5),
        (50_000_000_000, 50.0),
        (123_456_789, 0.12),  # 2-decimal rounding
        (12_345_678_901, 12.35),
    ],
)
def test_bytes_to_gb(bytes_in, gb_out):
    assert _bytes_to_gb(bytes_in) == gb_out


@pytest.mark.parametrize(
    "ms_in,sec_out",
    [
        (None, None),
        (0, 0.0),
        (1000, 1.0),
        (12_345, 12.35),  # 2-decimal rounding
        (123, 0.12),
        (3_600_000, 3600.0),  # 1 hour in ms
    ],
)
def test_ms_to_sec(ms_in, sec_out):
    assert _ms_to_sec(ms_in) == sec_out


# ---- COLUMNS shape ----

def test_columns_have_expected_display_names():
    """The Excel header row uses the display names — these are the
    user-visible names, including the unit-suffixed ones (sec, gb)."""
    display_names = [c[0] for c in COLUMNS]
    assert display_names == [
        "query_id",
        "sql",
        "elapsed_sec",
        "gb_scanned",
        "estimated_credits",
        "warehouse_name",
        "warehouse_size",
        "recommendation_type",
        "recommendation_target",
        "rationale",
    ]
    # The OLD optimized/delta/validation column names must NOT appear —
    # they belonged to the rewrite-and-validate flow that the analyzer
    # replaced.
    for legacy in (
        "original_sql",
        "optimized_sql",
        "validation_status",
        "original_elapsed_sec",
        "optimized_elapsed_sec",
        "elapsed_delta_pct",
        "original_gb_scanned",
        "optimized_gb_scanned",
        "bytes_delta_pct",
        "original_credits",
        "optimized_credits",
        "credits_delta_pct",
        "pattern_applied",
        "recommended_warehouse_size",
        "notes",
    ):
        assert legacy not in display_names


def test_columns_use_correct_source_keys():
    """Records still come in with raw bytes/ms keys — the writer reads
    those source keys, applies conversion, and writes the converted value
    under the unit-suffixed display name. If a source key gets renamed,
    every existing optimizations.json would silently produce blank cells."""
    by_display = {c[0]: c[1] for c in COLUMNS}
    assert by_display["elapsed_sec"] == "elapsed_ms"
    assert by_display["gb_scanned"] == "bytes_scanned"
    # Pass-through columns map display name 1:1 to source key.
    assert by_display["query_id"] == "query_id"
    assert by_display["sql"] == "sql"
    assert by_display["recommendation_type"] == "recommendation_type"
    assert by_display["recommendation_target"] == "recommendation_target"
    assert by_display["rationale"] == "rationale"


# ---- end-to-end: write + read back ----

def _write_and_read(records: list[dict]):
    path = tempfile.mktemp(suffix=".xlsx")
    try:
        write_report(records, path)
        wb = load_workbook(path)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append({headers[i - 1]: ws.cell(row=r, column=i).value for i in range(1, ws.max_column + 1)})
        return headers, rows
    finally:
        if os.path.exists(path):
            os.remove(path)


def test_write_report_converts_bytes_to_gb_and_ms_to_sec():
    """The crucial assertion: the user opens the spreadsheet and sees GB
    and seconds, not bytes and ms."""
    rec = {
        "query_id": "01a",
        "sql": "SELECT * FROM foo",
        "elapsed_ms": 12_500,            # → 12.5 sec
        "bytes_scanned": 5_000_000_000,  # → 5.0 GB
        "estimated_credits": 0.42,
        "warehouse_name": "OPTIMIZE_WH",
        "warehouse_size": "Medium",
        "recommendation_type": "search-optimization",
        "recommendation_target": "FOO.X",
        "rationale": "TableScan with equality filter on X",
    }
    _, rows = _write_and_read([rec])
    assert rows[0]["elapsed_sec"] == 12.5
    assert rows[0]["gb_scanned"] == 5.0


def test_write_report_handles_null_unit_fields():
    """Some candidates may have null bytes_scanned (e.g. if QUERY_HISTORY
    didn't yet have the row when discovery wrote candidates.json). They
    should write as blank cells, not crash on the conversion."""
    rec = {
        "query_id": "02b",
        "sql": "SELECT 1",
        "elapsed_ms": 8_000,
        "bytes_scanned": None,
        "estimated_credits": None,
        "warehouse_name": "OPTIMIZE_WH",
        "warehouse_size": "Small",
        "recommendation_type": "none",
        "recommendation_target": None,
        "rationale": "No clear pattern match.",
    }
    _, rows = _write_and_read([rec])
    assert rows[0]["elapsed_sec"] == 8.0
    assert rows[0]["gb_scanned"] is None
    assert rows[0]["estimated_credits"] is None
    assert rows[0]["recommendation_target"] is None


def test_write_report_emits_recommendation_columns():
    """The two columns that carry the analyzer's actionable output must
    round-trip cleanly — they're the whole point of the report."""
    rec = {
        "query_id": "03c",
        "sql": "SELECT count(*) FROM events WHERE event_date >= ?",
        "elapsed_ms": 30_000,
        "bytes_scanned": 8_000_000_000,
        "estimated_credits": 1.5,
        "warehouse_name": "OPTIMIZE_WH",
        "warehouse_size": "Large",
        "recommendation_type": "clustering-key",
        "recommendation_target": "EVENTS(EVENT_DATE)",
        "rationale": "950/1000 partitions scanned despite selective date filter.",
    }
    headers, rows = _write_and_read([rec])
    assert "recommendation_type" in headers
    assert "recommendation_target" in headers
    assert rows[0]["recommendation_type"] == "clustering-key"
    assert rows[0]["recommendation_target"] == "EVENTS(EVENT_DATE)"


def test_write_report_truncates_very_long_sql():
    long_sql = "SELECT " + ("x, " * 4000) + "y FROM t"
    rec = {
        "query_id": "04d",
        "sql": long_sql,
    }
    _, rows = _write_and_read([rec])
    assert rows[0]["sql"].endswith("…")
    assert len(rows[0]["sql"]) <= 8000
